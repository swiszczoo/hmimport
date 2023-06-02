from datetime import datetime
from openpyxl import Workbook, load_workbook
from openpyxl.cell.cell import Cell
from openpyxl.worksheet.worksheet import Worksheet
from sys import argv, stderr
from typing import Any, Optional


PK_PREFIX = 'Id'

def get_column_name(index: int) -> str:
    return chr(ord('A') + index)

def recognize_type(value: Any) -> Optional[str]:
    if value is None:
        return None
    if type(value) == str:
        return 'VARCHAR(255)'
    if type(value) == int:
        return 'INTEGER'
    if type(value) == float:
        return 'DOUBLE'
    if type(value) == datetime:
        if value.hour == 0 and value.minute == 0 and value.second == 0:
            return 'DATE'
        else:
            return 'DATETIME'
        
    raise (Exception('unknown data type'))

def can_promote_type(frm: str, to: str) -> bool:
    types = ['INTEGER', 'DOUBLE', 'DATE', 'DATETIME', 'VARCHAR(255)']
    return frm is None or types.index(frm) <= types.index(to)

def determine_column_names(sheet: Worksheet) -> list[str]:
    return list(map(lambda cell: cell.value, sheet[1]))

def determine_data_types(sheet: Worksheet) -> list[str]:
    output = []
    for i, column in enumerate(sheet[1]):
        optional = False
        type = None
        for cell in sheet[get_column_name(i)][1:]:
            this_type = recognize_type(cell.value)
            if this_type is None:
                optional = True
            else:
                if type != this_type:
                    if can_promote_type(type, this_type):
                        type = this_type
                
        if type == None:
            raise(Exception('empty column found'))
        
        output.append(type + ('' if optional else ' NOT NULL'))
    return output

def build_create_table_query(table_name: str, column_names: list[str], data_types: list[str]) -> str:
    pk = []
    for column_name in column_names:
        if column_name.startswith(PK_PREFIX):
            pk.append(column_name)
        else:
            break

    if not pk:
        pk.append(column_names[0])

    def column_def(data: tuple[str, str]) -> str:
        attributes = ''
        if len(pk) == 1 and data[0].startswith(PK_PREFIX):
            attributes = ' AUTO_INCREMENT'
        return f'    {data[0]} {data[1]}{attributes}'
    
    return (
        f'CREATE TABLE `{table_name}` (\n' 
        + ',\n'.join(map(column_def, zip(column_names, data_types)))
        + ',\n    PRIMARY KEY(' + ', '.join(pk) + ')\n);'
    )

def build_data_insert_query(table_name: str, sheet: Worksheet, column_names: list[str]) -> str:
    def value_to_str(value: Any):
        if type(value) == datetime:
            return f"'{value.strftime('%Y-%m-%d')}'"
        elif type(value) == float:
            return f'{value:.2f}'
        elif value is None:
            return 'NULL'
        else:
            return repr(value)
        
    def row_tuple(row: tuple[Cell, ...]):
        return '(' + ', '.join(map(value_to_str, map(lambda cell: cell.value, row))) + ')'

    skipped_header = sheet.__iter__()
    next(skipped_header)

    return (
        f'INSERT INTO `{table_name}` (' + ', '.join(column_names) + ') VALUES\n'
        + ',\n'.join(map(row_tuple, skipped_header)) + ';'
    )

if __name__ == '__main__':
    if len(argv) != 2:
        stderr.write('Usage: python hmimport.py [xlsx_file]\n')
        exit(1)

    try:
        wb: Workbook = load_workbook(argv[1])
    except:
        stderr.write('Couldn\'t read xlsx workbook!')
        exit(2)

    if len(wb.sheetnames) != 1:
        stderr.write('Couldn\'t read workbook that contains more than one worksheet')

    name: str = wb.sheetnames[0]
    sheet: Worksheet = wb[name]
    columns: list[str] = determine_column_names(sheet)
    data_types: list[str] = determine_data_types(sheet)
    print(build_create_table_query(name, columns, data_types))
    print(build_data_insert_query(name, sheet, columns))
    print('')
    